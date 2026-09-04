from pathlib import Path


import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

def create_excel_report(
    risk_report,
    risk_contributions,
    rolling_volatility,
    rolling_tracking_error,
    portfolio_returns,
    benchmark_returns,
    output_path,
):
    """
    Create an Excel portfolio risk report.
    """

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()

    # Remove default worksheet
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    # --------------------------------------------------
    # Executive Summary
    # --------------------------------------------------

    summary = workbook.create_sheet("Executive Summary")

    summary["A1"] = "Portfolio Risk Report"
    summary["A1"].font = Font(bold=True, size=16)

    summary["A3"] = "Metric"
    summary["B3"] = "Value"

    summary["A3"].font = Font(bold=True)
    summary["B3"].font = Font(bold=True)

    summary_metrics = [
        "Annualised Volatility",
        "Maximum Drawdown",
        "95% VaR",
        "99% VaR",
        "95% Expected Shortfall",
        "Tracking Error",
        "Information Ratio",
    ]

    for row, metric in enumerate(summary_metrics, start=4):
        summary.cell(row=row, column=1, value=metric)
        summary.cell(row=row, column=2, value=risk_report[metric])

    # --------------------------------------------------
    # Risk Metrics
    # --------------------------------------------------

    metrics_sheet = workbook.create_sheet("Risk Metrics")

    metrics_sheet["A1"] = "Risk Metrics"
    metrics_sheet["A1"].font = Font(bold=True, size=16)

    metrics_sheet["A3"] = "Metric"
    metrics_sheet["B3"] = "Value"

    metrics_sheet["A3"].font = Font(bold=True)
    metrics_sheet["B3"].font = Font(bold=True)

    for row, (metric, value) in enumerate(risk_report.items(), start=4):
        metrics_sheet.cell(row=row, column=1, value=metric)
        metrics_sheet.cell(row=row, column=2, value=float(value))

    # --------------------------------------------------
    # Risk Contributions
    # --------------------------------------------------

    contributions_sheet = workbook.create_sheet("Risk Contributions")

    contributions_sheet["A1"] = "Portfolio Risk Contributions"
    contributions_sheet["A1"].font = Font(bold=True, size=16)

    contributions_df = risk_contributions.reset_index()

    for column_index, column_name in enumerate(
        contributions_df.columns,
        start=1,
    ):
        contributions_sheet.cell(
            row=3,
            column=column_index,
            value=column_name,
        )
        contributions_sheet.cell(
            row=3,
            column=column_index,
        ).font = Font(bold=True)

    for row_index, row in enumerate(
        contributions_df.itertuples(index=False),
        start=4,
    ):
        for column_index, value in enumerate(row, start=1):
            contributions_sheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

    # --------------------------------------------------
    # Benchmark Relative
    # --------------------------------------------------

    benchmark_sheet = workbook.create_sheet("Benchmark Relative")

    benchmark_sheet["A1"] = "Portfolio vs Benchmark"
    benchmark_sheet["A1"].font = Font(bold=True, size=16)

    benchmark_df = pd.DataFrame(
        {
            "Portfolio Return": portfolio_returns,
            "Benchmark Return": benchmark_returns,
            "Active Return": portfolio_returns - benchmark_returns,
        }
    )

    benchmark_df = benchmark_df.reset_index()

    for column_index, column_name in enumerate(
        benchmark_df.columns,
        start=1,
    ):
        benchmark_sheet.cell(
            row=3,
            column=column_index,
            value=column_name,
        )
        benchmark_sheet.cell(
            row=3,
            column=column_index,
        ).font = Font(bold=True)

    for row_index, row in enumerate(
        benchmark_df.itertuples(index=False),
        start=4,
    ):
        for column_index, value in enumerate(row, start=1):
            benchmark_sheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

    # --------------------------------------------------
    # Rolling Risk
    # --------------------------------------------------

    rolling_sheet = workbook.create_sheet("Rolling Risk")

    rolling_sheet["A1"] = "Rolling Risk Measures"
    rolling_sheet["A1"].font = Font(bold=True, size=16)

    rolling_df = pd.DataFrame(
        {
            "Rolling Volatility": rolling_volatility,
            "Rolling Tracking Error": rolling_tracking_error,
        }
    )

    rolling_df = rolling_df.reset_index()

    for column_index, column_name in enumerate(
        rolling_df.columns,
        start=1,
    ):
        rolling_sheet.cell(
            row=3,
            column=column_index,
            value=column_name,
        )
        rolling_sheet.cell(
            row=3,
            column=column_index,
        ).font = Font(bold=True)

    for row_index, row in enumerate(
        rolling_df.itertuples(index=False),
        start=4,
    ):
        for column_index, value in enumerate(row, start=1):
            rolling_sheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

    # --------------------------------------------------
    # Professional formatting
    # --------------------------------------------------




    percentage_metrics = {
        "Annualised Volatility",
        "Maximum Drawdown",
        "95% VaR",
        "99% VaR",
        "95% Expected Shortfall",
        "Tracking Error",
    }

    dollar_metrics = {
        "95% VaR ($)",
        "99% VaR ($)",
        "95% Expected Shortfall ($)",
    }


    def format_metric_sheet(worksheet):
        worksheet.freeze_panes = "A4"

        worksheet.column_dimensions["A"].width = 32
        worksheet.column_dimensions["B"].width = 20

        for row in range(4, worksheet.max_row + 1):
            metric = worksheet.cell(row=row, column=1).value
            value_cell = worksheet.cell(row=row, column=2)

            if metric in percentage_metrics:
                value_cell.number_format = "0.00%"

            elif metric in dollar_metrics:
                value_cell.number_format = '$#,##0.00'

            else:
                value_cell.number_format = "0.0000"

        for cell in worksheet[3]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        worksheet.auto_filter.ref = (
            f"A3:B{worksheet.max_row}"
        )


    def format_data_sheet(worksheet):
        worksheet.freeze_panes = "A4"

        for column_cells in worksheet.columns:
            max_length = 0

            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(
                        max_length,
                        len(str(cell.value)),
                    )

            column_letter = get_column_letter(
                column_cells[0].column
            )

            worksheet.column_dimensions[
                column_letter
            ].width = min(max_length + 3, 35)

        for cell in worksheet[3]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        worksheet.auto_filter.ref = (
            f"A3:{get_column_letter(worksheet.max_column)}"
            f"{worksheet.max_row}"
        )


    # Executive Summary
    summary.freeze_panes = "A4"
    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 20

    for row in range(4, summary.max_row + 1):
        metric = summary.cell(row=row, column=1).value
        value_cell = summary.cell(row=row, column=2)

        if metric in percentage_metrics:
            value_cell.number_format = "0.00%"

        elif metric in dollar_metrics:
            value_cell.number_format = '$#,##0.00'

        elif metric == "Information Ratio":
            value_cell.number_format = "0.00"

        else:
            value_cell.number_format = "0.0000"

    for cell in summary[3]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")


    # Risk Metrics
    format_metric_sheet(metrics_sheet)


    # Risk Contributions
    format_data_sheet(contributions_sheet)

    for row in range(4, contributions_sheet.max_row + 1):
        for column in range(2, contributions_sheet.max_column + 1):
            contributions_sheet.cell(
                row=row,
                column=column,
            ).number_format = "0.00%"


    # Benchmark Relative
    format_data_sheet(benchmark_sheet)

    for row in range(4, benchmark_sheet.max_row + 1):
        for column in range(2, benchmark_sheet.max_column + 1):
            benchmark_sheet.cell(
                row=row,
                column=column,
            ).number_format = "0.00%"


    # Rolling Risk
    format_data_sheet(rolling_sheet)

    for row in range(4, rolling_sheet.max_row + 1):
        for column in range(2, rolling_sheet.max_column + 1):
            rolling_sheet.cell(
                row=row,
                column=column,
            ).number_format = "0.00%"


    # Conditional formatting for risk contributions
    if contributions_sheet.max_row >= 4:
        contributions_sheet.conditional_formatting.add(
            f"B4:B{contributions_sheet.max_row}",
            ColorScaleRule(
                start_type="min",
                start_color="FFFFFF",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFFF00",
                end_type="max",
                end_color="FF0000",
            ),
        )


    # Conditional formatting for rolling volatility
    if rolling_sheet.max_row >= 4:
        rolling_sheet.conditional_formatting.add(
            f"B4:B{rolling_sheet.max_row}",
            ColorScaleRule(
                start_type="min",
                start_color="FFFFFF",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFFF00",
                end_type="max",
                end_color="FF0000",
            ),
        )


    workbook.save(output_path)

    workbook.save(output_path)